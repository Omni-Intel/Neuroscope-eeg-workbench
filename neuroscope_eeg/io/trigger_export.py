from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
import unicodedata

import numpy as np

from neuroscope_eeg.timing.codebook import CODEBOOK_VERSION, EVENT_CODES
from neuroscope_eeg.timing.models import HardwareTriggerSample, TriggerDispatch


_HEADER_FILL = "0F766E"
_HEADER_FONT = "FFFFFF"
_SUBTLE_FILL = "E6FFFA"
_WARNING_FILL = "FEF3C7"
_ERROR_FILL = "FEE2E2"


def _hardware_lsl_delta_ms(dispatch: TriggerDispatch) -> tuple[float | None, float | None]:
    bridge = dispatch.clock_bridge
    if (
        bridge is None
        or dispatch.hardware_dispatch_time is None
        or dispatch.lsl_timestamp is None
    ):
        return None, None
    hardware_lsl_time = dispatch.hardware_dispatch_time + bridge.lsl_minus_monotonic
    return (dispatch.lsl_timestamp - hardware_lsl_time) * 1000.0, bridge.uncertainty_ms


def _base_row(dispatch: TriggerDispatch) -> dict[str, Any]:
    delta_ms, uncertainty_ms = _hardware_lsl_delta_ms(dispatch)
    return {
        "event_id": dispatch.event_id,
        "sequence": dispatch.sequence,
        "hardware_code": dispatch.hardware_code,
        "hardware_symbol": dispatch.hardware_symbol,
        "paradigm": dispatch.paradigm,
        "phase": dispatch.phase,
        "label": dispatch.label,
        "condition": dispatch.payload.get("condition"),
        "block": dispatch.payload.get("block_index", dispatch.payload.get("block")),
        "trial": dispatch.payload.get("trial_index", dispatch.payload.get("trial")),
        "wall_time": dispatch.wall_time,
        "wall_time_iso": (
            datetime.fromtimestamp(dispatch.wall_time, timezone.utc).isoformat()
            if dispatch.wall_time > 0
            else None
        ),
        "intent_time": dispatch.intent_time,
        "onset_hook_time": dispatch.onset_hook_time,
        "hook_type": dispatch.hook_type,
        "hardware_dispatch_time": dispatch.hardware_dispatch_time,
        "hardware_write_complete_time": dispatch.hardware_write_complete_time,
        "lsl_timestamp": dispatch.lsl_timestamp,
        "hardware_lsl_delta_ms": delta_ms,
        "conversion_uncertainty_ms": uncertainty_ms,
        "hardware_error": dispatch.hardware_error,
        "lsl_error": dispatch.lsl_error,
        "hardware_source_sample_index": None,
        "hardware_sample_index": None,
        "eeg_time_sec": None,
        "physical_onset_sample_index": None,
        "physical_onset_calibrated": False,
        "pairing_status": "not_requested" if not dispatch.hardware_requested else "pending",
        "timing_status": dispatch.timing_status,
        "payload": dict(dispatch.payload),
    }


def _percentile(values: list[float], percentile: float) -> float | None:
    return None if not values else float(np.percentile(np.asarray(values, dtype=np.float64), percentile))


def pair_trigger_events(
    dispatches: Iterable[TriggerDispatch],
    hardware_samples: Iterable[HardwareTriggerSample],
    *,
    sfreq: float,
    source_sample_offset: int = 0,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if sfreq <= 0:
        raise ValueError("sfreq must be positive")
    ordered_dispatches = sorted(dispatches, key=lambda item: item.sequence)
    samples = sorted(hardware_samples, key=lambda item: item.sample_index)
    rows: list[dict[str, Any]] = []
    cursor = 0
    unexpected = 0
    for dispatch in ordered_dispatches:
        row = _base_row(dispatch)
        if not dispatch.hardware_requested:
            rows.append(row)
            continue
        if dispatch.hardware_error or dispatch.hardware_code is None:
            row["pairing_status"] = "send_failed"
            rows.append(row)
            continue
        matching_index = next(
            (
                index
                for index in range(cursor, len(samples))
                if samples[index].code == dispatch.hardware_code
            ),
            None,
        )
        if matching_index is None:
            row["pairing_status"] = "missing_hardware"
            rows.append(row)
            continue
        if matching_index == cursor:
            row["pairing_status"] = "matched"
        else:
            row["pairing_status"] = "out_of_order"
            unexpected += matching_index - cursor
        sample = samples[matching_index]
        recording_sample_index = sample.sample_index - int(source_sample_offset)
        row["hardware_source_sample_index"] = sample.sample_index
        row["hardware_sample_index"] = recording_sample_index
        row["eeg_time_sec"] = recording_sample_index / float(sfreq)
        row["hardware_channel"] = sample.channel_name
        if row["pairing_status"] == "matched":
            row["timing_status"] = "hardware_sample_locked"
        cursor = matching_index + 1
        rows.append(row)
    unexpected += max(0, len(samples) - cursor)
    deltas = [
        float(row["hardware_lsl_delta_ms"])
        for row in rows
        if row["hardware_lsl_delta_ms"] is not None
    ]
    summary = {
        "codebook_version": CODEBOOK_VERSION,
        "total_events": len(rows),
        "hardware_requested_events": sum(bool(item.hardware_requested) for item in ordered_dispatches),
        "hardware_samples_received": len(samples),
        "source_sample_offset": int(source_sample_offset),
        "matched_events": sum(row["pairing_status"] == "matched" for row in rows),
        "out_of_order_events": sum(row["pairing_status"] == "out_of_order" for row in rows),
        "missing_hardware_events": sum(row["pairing_status"] == "missing_hardware" for row in rows),
        "send_failed_events": sum(row["pairing_status"] == "send_failed" for row in rows),
        "unexpected_hardware_events": unexpected,
        "lsl_failed_events": sum(bool(item.lsl_error) for item in ordered_dispatches),
        "hardware_lsl_delta_ms": {
            "median": _percentile(deltas, 50),
            "p95": _percentile(deltas, 95),
            "p99": _percentile(deltas, 99),
            "max": max(deltas) if deltas else None,
        },
    }
    return rows, summary


def _openpyxl() -> tuple[Any, Any, Any, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("Excel 导出需要 openpyxl>=3.1") from exc
    return Workbook, Alignment, Font, PatternFill


def _format_table_sheet(sheet: Any, *, status_column: int | None = None) -> None:
    _Workbook, Alignment, Font, PatternFill = _openpyxl()
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.font = Font(color=_HEADER_FONT, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 26
    for column_cells in sheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        display_widths = [
            sum(2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1 for character in value)
            for value in values
        ]
        width = min(42, max(10, max(display_widths) + 2))
        sheet.column_dimensions[column_cells[0].column_letter].width = width
    if status_column is not None:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=status_column)
            value = str(cell.value or "")
            if value in {"matched", "hardware_sample_locked"}:
                cell.fill = PatternFill("solid", fgColor=_SUBTLE_FILL)
            elif value in {"missing_hardware", "send_failed", "hardware_failed", "unsynchronized"}:
                cell.fill = PatternFill("solid", fgColor=_ERROR_FILL)
            elif value not in {"", "not_requested"}:
                cell.fill = PatternFill("solid", fgColor=_WARNING_FILL)


def _write_codebook_workbook(path: Path) -> None:
    Workbook, _Alignment, _Font, _PatternFill = _openpyxl()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "事件码对照"
    headers = (
        "硬件码",
        "符号名",
        "范式",
        "阶段",
        "条件",
        "中文含义",
        "DCP命令类型",
        "关键事件",
        "适用模式",
        "码表版本",
    )
    sheet.append(headers)
    for definition in EVENT_CODES:
        sheet.append(
            (
                definition.code,
                definition.symbol,
                definition.paradigm,
                definition.phase,
                definition.condition,
                definition.description_zh,
                definition.dcp_command,
                "是" if definition.critical else "否",
                "hardware_lsl / lsl_only",
                CODEBOOK_VERSION,
            )
        )
    _format_table_sheet(sheet)
    for cell in sheet["A"][1:]:
        cell.number_format = "0"
    workbook.save(path)


_TIMELINE_COLUMNS: tuple[tuple[str, str], ...] = (
    ("事件编号", "event_id"),
    ("顺序号", "sequence"),
    ("硬件码", "hardware_code"),
    ("符号名", "hardware_symbol"),
    ("范式", "paradigm"),
    ("阶段", "phase"),
    ("标签", "label"),
    ("条件", "condition"),
    ("Block", "block"),
    ("Trial", "trial"),
    ("本机墙钟Unix时间", "wall_time"),
    ("本机墙钟UTC", "wall_time_iso"),
    ("Intent时间", "intent_time"),
    ("Onset Hook时间", "onset_hook_time"),
    ("Hook类型", "hook_type"),
    ("DCP发送时间", "hardware_dispatch_time"),
    ("DCP完成时间", "hardware_write_complete_time"),
    ("LSL时间戳", "lsl_timestamp"),
    ("采集源硬件采样点", "hardware_source_sample_index"),
    ("BDF内硬件采样点", "hardware_sample_index"),
    ("BDF内EEG时间/s", "eeg_time_sec"),
    ("硬件-LSL差/ms", "hardware_lsl_delta_ms"),
    ("换算不确定度/ms", "conversion_uncertainty_ms"),
    ("配对状态", "pairing_status"),
    ("同步状态", "timing_status"),
    ("硬件错误", "hardware_error"),
    ("LSL错误", "lsl_error"),
    ("物理起始采样点", "physical_onset_sample_index"),
    ("物理校准", "physical_onset_calibrated"),
)


def _write_timeline_workbook(
    path: Path,
    rows: list[dict[str, Any]],
    summary: dict[str, Any],
    timing_config: dict[str, Any],
) -> None:
    Workbook, Alignment, Font, PatternFill = _openpyxl()
    workbook = Workbook()
    timeline = workbook.active
    timeline.title = "事件时间线"
    timeline.append(tuple(label for label, _key in _TIMELINE_COLUMNS))
    for row in rows:
        timeline.append(tuple(row.get(key) for _label, key in _TIMELINE_COLUMNS))
    pairing_column = next(index for index, item in enumerate(_TIMELINE_COLUMNS, start=1) if item[1] == "pairing_status")
    _format_table_sheet(timeline, status_column=pairing_column)
    numeric_six_decimals = {
        "wall_time",
        "intent_time",
        "onset_hook_time",
        "hardware_dispatch_time",
        "hardware_write_complete_time",
        "lsl_timestamp",
        "eeg_time_sec",
    }
    numeric_three_decimals = {"hardware_lsl_delta_ms", "conversion_uncertainty_ms"}
    for index, (_label, key) in enumerate(_TIMELINE_COLUMNS, start=1):
        if key in numeric_six_decimals:
            for column in timeline.iter_cols(min_col=index, max_col=index, min_row=2):
                for cell in column:
                    cell.number_format = "0.000000"
        elif key in numeric_three_decimals:
            for column in timeline.iter_cols(min_col=index, max_col=index, min_row=2):
                for cell in column:
                    cell.number_format = "0.000"

    summary_sheet = workbook.create_sheet("会话同步摘要")
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.append(("项目", "值", "说明"))
    summary_rows = (
        ("运行模式", timing_config.get("mode", ""), "hardware_lsl 或 lsl_only"),
        ("总体同步状态", summary["overall_timing_status"], "只有实际配对后才是 hardware_sample_locked"),
        ("NDE0001串口", timing_config.get("port", ""), "仅硬件模式使用"),
        ("EEG采样率/Hz", timing_config.get("sfreq", ""), "硬件采样点换算依据"),
        ("BDF起始源采样点", timing_config.get("source_sample_offset", 0), "采集源采样点减此值即 BDF 内采样点"),
        ("物理起始已校准", bool(timing_config.get("physical_onset_calibrated", False)), "现场无光电/音频回环时为 FALSE"),
        ("事件总数", summary["total_events"], "包括仅 LSL 元数据事件"),
        ("请求硬件事件数", summary["hardware_requested_events"], "应发送 NDE0001 DCP 的事件"),
        ("收到硬件样本数", summary["hardware_samples_received"], "JellyFish Trigger/Event 通道边沿"),
        ("完全匹配", summary["matched_events"], "码值和顺序一致"),
        ("乱序匹配", summary["out_of_order_events"], "找到同码但之前存在额外硬件事件"),
        ("缺失硬件事件", summary["missing_hardware_events"], "DCP 写入日志存在但 EEG Trigger 未找到"),
        ("意外硬件事件", summary["unexpected_hardware_events"], "EEG Trigger 存在但没有对应本地事件"),
        ("硬件发送失败", summary["send_failed_events"], "DCP 写入失败"),
        ("LSL发送失败", summary["lsl_failed_events"], "LSL Marker 推送失败"),
        ("硬件-LSL差中位数/ms", summary["hardware_lsl_delta_ms"]["median"], "经时钟桥换算"),
        ("硬件-LSL差P95/ms", summary["hardware_lsl_delta_ms"]["p95"], "经时钟桥换算"),
        ("硬件-LSL差P99/ms", summary["hardware_lsl_delta_ms"]["p99"], "经时钟桥换算"),
        ("物理起始状态", "未校准", "未接光电或音频回环，不填物理起始时间"),
    )
    for item in summary_rows:
        summary_sheet.append(item)
    _format_table_sheet(summary_sheet)
    summary_sheet.column_dimensions["A"].width = 24
    summary_sheet.column_dimensions["B"].width = 22
    summary_sheet.column_dimensions["C"].width = 54
    for cell in summary_sheet["C"][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in summary_sheet["A"][1:]:
        cell.font = Font(bold=True)
    for row in range(2, summary_sheet.max_row + 1):
        if summary_sheet.cell(row=row, column=1).value == "物理起始状态":
            summary_sheet.cell(row=row, column=2).fill = PatternFill("solid", fgColor=_WARNING_FILL)
    workbook.save(path)


def export_trigger_artifacts(
    session_dir: Path | str,
    dispatches: Iterable[TriggerDispatch],
    hardware_samples: Iterable[HardwareTriggerSample],
    *,
    sfreq: float,
    timing_config: dict[str, Any],
    source_sample_offset: int = 0,
) -> dict[str, Any]:
    root = Path(session_dir)
    root.mkdir(parents=True, exist_ok=True)
    dispatch_list = list(dispatches)
    sample_list = list(hardware_samples)
    rows, summary = pair_trigger_events(
        dispatch_list,
        sample_list,
        sfreq=sfreq,
        source_sample_offset=source_sample_offset,
    )
    config = {
        **timing_config,
        "sfreq": float(sfreq),
        "source_sample_offset": int(source_sample_offset),
    }
    summary_payload = {
        **summary,
        "timing_mode": config.get("mode", ""),
        "triggerbox_port": config.get("port", ""),
        "physical_onset_calibrated": bool(config.get("physical_onset_calibrated", False)),
        "physical_onset_status": "uncalibrated",
    }
    if summary_payload["timing_mode"] == "lsl_only":
        overall_status = "lsl_software_sync_uncalibrated"
    elif summary_payload["send_failed_events"]:
        overall_status = "hardware_failed"
    elif (
        summary_payload["hardware_requested_events"] > 0
        and summary_payload["matched_events"] == summary_payload["hardware_requested_events"]
    ):
        overall_status = "hardware_sample_locked"
    else:
        overall_status = "hardware_dispatched_unverified"
    summary_payload["overall_timing_status"] = overall_status
    (root / "event_codebook.json").write_text(
        json.dumps(
            {
                "codebook_version": CODEBOOK_VERSION,
                "events": [definition.as_dict() for definition in EVENT_CODES],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    (root / "synchronization_summary.json").write_text(
        json.dumps(summary_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    _write_codebook_workbook(root / "event_codebook.xlsx")
    _write_timeline_workbook(root / "event_timeline.xlsx", rows, summary_payload, config)
    return {"rows": rows, "summary": summary_payload}
