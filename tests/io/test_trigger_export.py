from __future__ import annotations

import json

from openpyxl import load_workbook

from neuroscope_eeg.io.trigger_export import export_trigger_artifacts, pair_trigger_events
from neuroscope_eeg.timing.models import (
    ClockBridgeSample,
    HardwareTriggerSample,
    TriggerDispatch,
)


def dispatch(sequence: int, code: int, *, requested: bool = True) -> TriggerDispatch:
    return TriggerDispatch(
        event_id=f"EVT-{sequence:06d}",
        sequence=sequence,
        session_id="session-1",
        paradigm="N-back 工作记忆",
        phase="nback_trial",
        label="7",
        payload={"block_index": 2, "trial_index": sequence - 1, "condition": "target"},
        wall_time=1_786_400_000.0 + sequence,
        intent_time=9.9 + sequence,
        onset_hook_time=10.0 + sequence,
        hook_type="frame_swapped",
        timing_mode="hardware_lsl",
        timing_status="hardware_dispatched_unverified",
        hardware_code=code,
        hardware_symbol="NBACK_1_TARGET",
        hardware_requested=requested,
        hardware_frame_hex=f"01 e1 01 00 {code:02x}",
        hardware_dispatch_time=10.0 + sequence,
        hardware_write_complete_time=10.001 + sequence,
        lsl_timestamp=110.002 + sequence,
        clock_bridge=ClockBridgeSample(9.999 + sequence, 110.0 + sequence, 10.001 + sequence),
    )


def test_pairing_only_claims_sample_lock_for_real_neuracle_trigger_samples() -> None:
    dispatches = [dispatch(1, 53), dispatch(2, 55), dispatch(3, 51)]
    samples = [
        HardwareTriggerSample(53, 1000, "TRIGGER"),
        HardwareTriggerSample(99, 1500, "TRIGGER"),
        HardwareTriggerSample(55, 2000, "TRIGGER"),
    ]
    rows, summary = pair_trigger_events(
        dispatches,
        samples,
        sfreq=1000.0,
        source_sample_offset=100,
    )
    assert rows[0]["pairing_status"] == "matched"
    assert rows[0]["hardware_source_sample_index"] == 1000
    assert rows[0]["hardware_sample_index"] == 900
    assert rows[0]["eeg_time_sec"] == 0.9
    assert rows[0]["timing_status"] == "hardware_sample_locked"
    assert rows[1]["pairing_status"] == "out_of_order"
    assert rows[1]["hardware_source_sample_index"] == 2000
    assert rows[1]["hardware_sample_index"] == 1900
    assert rows[2]["pairing_status"] == "missing_hardware"
    assert rows[2]["hardware_sample_index"] is None
    assert rows[2]["timing_status"] == "hardware_dispatched_unverified"
    assert summary["matched_events"] == 1
    assert summary["out_of_order_events"] == 1
    assert summary["missing_hardware_events"] == 1
    assert summary["unexpected_hardware_events"] == 1


def test_export_writes_clear_chinese_workbooks_and_machine_readable_summary(tmp_path) -> None:
    dispatches = [dispatch(1, 53), dispatch(2, 55)]
    samples = [HardwareTriggerSample(53, 1000, "TRIGGER"), HardwareTriggerSample(55, 2000, "TRIGGER")]
    result = export_trigger_artifacts(
        tmp_path,
        dispatches,
        samples,
        sfreq=1000.0,
        timing_config={"mode": "hardware_lsl", "port": "COM7", "physical_onset_calibrated": False},
    )
    assert result["summary"]["matched_events"] == 2
    assert (tmp_path / "event_codebook.json").is_file()
    assert (tmp_path / "event_codebook.xlsx").is_file()
    assert (tmp_path / "event_timeline.xlsx").is_file()
    summary = json.loads((tmp_path / "synchronization_summary.json").read_text(encoding="utf-8"))
    assert summary["timing_mode"] == "hardware_lsl"
    assert summary["physical_onset_calibrated"] is False

    codebook = load_workbook(tmp_path / "event_codebook.xlsx", data_only=False)
    assert codebook.sheetnames == ["事件码对照"]
    sheet = codebook["事件码对照"]
    assert sheet.freeze_panes == "A2"
    assert [cell.value for cell in sheet[1]][:4] == ["硬件码", "符号名", "范式", "阶段"]
    assert any(row[1].value == "NBACK_1_TARGET" for row in sheet.iter_rows(min_row=2))

    timeline = load_workbook(tmp_path / "event_timeline.xlsx", data_only=False)
    assert timeline.sheetnames == ["事件时间线", "会话同步摘要"]
    event_sheet = timeline["事件时间线"]
    headers = [cell.value for cell in event_sheet[1]]
    assert "采集源硬件采样点" in headers
    assert "BDF内硬件采样点" in headers
    assert "BDF内EEG时间/s" in headers
    assert "LSL时间戳" in headers
    assert "本机墙钟UTC" in headers
    assert "配对状态" in headers
    assert event_sheet.freeze_panes == "A2"
